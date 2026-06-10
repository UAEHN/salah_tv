"""
Rebuild Algeria prayer times CSV from 3 official Ministry xlsx files (1447H).
Covers: Adrar, Algiers, Djelfa base cities + their surrounding cities via offset tables.

Usage:
  python tool/rebuild_algeria_from_new_xlsx.py

Output: overwrites assets/csv/algeria_prayer_times_2026.csv (only the affected cities;
        all other cities in the file are kept unchanged).
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
from datetime import date, timedelta
import re, os, csv, copy

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(BASE_DIR, 'assets', 'csv', 'algeria_prayer_times_2026.csv')

# ── Hijri month normalization (strip tashkeel, variants) ─────────────────────
_TASHKEEL = re.compile(r'[ؐ-ًؚ-ٰٟۖ-ۜ۟-۪ۤۧۨ-ۭ]')
def clean_arabic(s):
    s = _TASHKEEL.sub('', s or '').strip()
    s = s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')  # alef variants
    return s

HIJRI_MONTHS_ORDER = [
    'محرم', 'صفر', 'ربيع الاول', 'ربيع الثاني',
    'جمادى الاولى', 'جمادى الثانية', 'رجب', 'شعبان',
    'رمضان', 'شوال', 'ذو القعدة', 'ذو الحجة',
]

def normalize_hijri(s):
    s = clean_arabic(s)
    # Common variant fixes
    s = s.replace('جمادي', 'جمادى').replace('ذو الحجه', 'ذو الحجة')
    s = s.replace('ربيع الأول', 'ربيع الاول').replace('ربيع الثانى', 'ربيع الثاني')
    s = s.replace('جمادى الأولى', 'جمادى الاولى').replace('جمادى الاولي', 'جمادى الاولى')
    s = s.replace('ذو القعده', 'ذو القعدة')
    return s

PRAYER_COLS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR_MAP = {
    'الفجر': 'fajr', 'الشروق': 'sunrise', 'الظهر': 'dhuhr',
    'العصر': 'asr', 'المغرب': 'maghrib', 'العشاء': 'isha',
}

# ── Arabic → English city name map ───────────────────────────────────────────

# Djelfa offset columns → CSV city names
DJELFA_AR_TO_EN = {
    'المغير':      'El Mghair',
    'أولاد جلال':  'Oulad Djallal',
    'مغنية':       'Maghnia',
    'سبدو':        'Sebdou',
    'عين توتة':    'Ain Temouchent',  # quirk: label "عين توتة" = عين تموشنت
    'تلمسان':      'Tlemcen',
    'ابن باديس':   'Ben Badis',
    'سيدي بلعباس': 'Sidi Bel Abbes',
    'الغَمَامة':   'Naama',           # quirk: "الغَمَامة" = النعامة
    'معسكر':       'Mascara',
    'سعيدة':       'Saida',
    'البيض':       'El Bayadh',
    'تيارت':       'Tiaret',
    'تيسمسيلت':   'Tissemsilt',
    'عين وسارة':   'Ain Wasara',
    'الأغواط':     'Al-Aghwat',
    'حاسي الرمل':  'Hassi Rmel',
    'عين الملح':   'Ain el Melh',
    'بوسعادة':     'Bousaada',
    'بسكرة':       'Biskra',
    'تقرت':        'Tuggurt',
    'باتنة':       'Batna',
    'الوادي':      'El Oued',
    'خنشلة':       'Khenchela',
    'بئر العاتر':  'Bir al-Ater',
    'تبسة':        'Tebessa',
}

# Algiers offset rows → CSV city names
ALGIERS_AR_TO_EN = {
    'وهران':       'Oran',
    'مستغانم':     'Mostaganem',
    'غليزان':      'Relizane',
    'الشلف':       'Chlef',
    'عين الدفلى':  'Ain Defla',
    'تيبازة':      'Tipaza',
    'المدية':      'Medea',
    'البليدة':     'Blida',
    'بومرداس':     'Boumerdes',
    'دلس':         'Dellys',
    'البويرة':     'Bouira',
    'تيزي وزو':   'Tizi Ouzou',
    'المسيلة':     'Msila',
    'برج بوعريريج':'Bordj Bou Arreridj',
    'بجاية':       'Bjaiya',
    'سطيف':        'Setif',
    'جيجل':        'Jijel',
    'ميلة':        'Milah',
    'قسنطينة':     'Constantine',
    'سكيكدة':      'Skikda',
    'أم البواقي':  'Um al-Bouaki',
    'قالمة':       'Qalma',
    'عنابة':       'Annaba',
    'سوق أهراس':   'Souk Ahras',
    'الطارف':      'El Tarf',
}

# ── Step 1: Read base xlsx files ──────────────────────────────────────────────

def parse_time(s):
    """'HH:MM' → minutes since midnight"""
    if not s or not isinstance(s, str):
        return None
    m = re.match(r'(\d{1,2}):(\d{2})', s.strip())
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))

def minutes_to_str(m):
    if m is None:
        return ''
    return f'{m // 60:02d}:{m % 60:02d}'

def read_base_xlsx(fpath):
    """
    Returns:
      times: dict {date_str_YYYYMMDD: {prayer: minutes}}   (354 rows)
      hijri: dict {date_str_YYYYMMDD: normalized_hijri_month}
    """
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    times = {}
    hijri = {}
    cur_month = None
    for row in ws.iter_rows(values_only=True):
        # Col B = date (format 2025/12/31 or datetime obj)
        date_val = row[1]
        if date_val is None:
            # Month header row: col A has month name
            if row[0]:
                cur_month = normalize_hijri(str(row[0]))
            continue
        # Parse date
        if isinstance(date_val, str) and re.match(r'\d{4}/\d{2}/\d{2}', date_val):
            parts = date_val.split('/')
            d = date(int(parts[0]), int(parts[1]), int(parts[2]))
        elif hasattr(date_val, 'year'):
            d = date_val if isinstance(date_val, date) else date_val.date()
        else:
            continue
        # Parse prayers: cols D-I (index 3-8)
        fajr    = parse_time(str(row[3]) if row[3] is not None else '')
        sunrise = parse_time(str(row[4]) if row[4] is not None else '')
        dhuhr   = parse_time(str(row[5]) if row[5] is not None else '')
        asr     = parse_time(str(row[6]) if row[6] is not None else '')
        maghrib = parse_time(str(row[7]) if row[7] is not None else '')
        isha    = parse_time(str(row[8]) if row[8] is not None else '')
        if fajr is None:
            continue
        key = d.strftime('%Y%m%d')
        times[key] = {'fajr': fajr, 'sunrise': sunrise, 'dhuhr': dhuhr,
                      'asr': asr, 'maghrib': maghrib, 'isha': isha}
        # Hijri month from col A of this row (or inherited from section header)
        if row[0] and str(row[0]).strip():
            cur_month = normalize_hijri(str(row[0]))
        if cur_month:
            hijri[key] = cur_month
    return times, hijri

print('Reading base xlsx files...')
adrar_times, adrar_hijri   = read_base_xlsx(os.path.join(BASE_DIR, 'adrar_new.xlsx'))
algiers_times, algiers_hijri = read_base_xlsx(os.path.join(BASE_DIR, 'algiers_new.xlsx'))
djelfa_times, djelfa_hijri  = read_base_xlsx(os.path.join(BASE_DIR, 'djelfa_new.xlsx'))

print(f'  Adrar:   {len(adrar_times)} days, Hijri months: {sorted(set(adrar_hijri.values()))}')
print(f'  Algiers: {len(algiers_times)} days, Hijri months: {sorted(set(algiers_hijri.values()))}')
print(f'  Djelfa:  {len(djelfa_times)} days, Hijri months: {sorted(set(djelfa_hijri.values()))}')

# ── Step 2: Build 2026 calendar mapping ──────────────────────────────────────

def build_2026_calendar(source_times, source_hijri):
    """
    Returns dict {date(2026,m,d): {'fajr':..., ..., 'hijri_month': str}}
    for all 365 days of 2026.

    Mapping:
      2026-01-01..2026-06-15 → use row with same date from file (2026 rows)
      2026-06-16..2026-06-26 → None (gap: keep current CSV values)
      2026-06-27..2026-12-31 → use row with same MM/DD from 2025 in file
    """
    result = {}
    d = date(2026, 1, 1)
    one_day = timedelta(days=1)
    end = date(2026, 12, 31)
    while d <= end:
        if d <= date(2026, 6, 15):
            key = d.strftime('%Y%m%d')  # 2026MMDD
        elif d <= date(2026, 6, 26):
            result[d] = None  # gap
            d += one_day
            continue
        else:
            # Use same MM/DD from 2025
            key = f'2025{d.month:02d}{d.day:02d}'
        row = source_times.get(key)
        if row:
            entry = dict(row)
            entry['hijri_month'] = source_hijri.get(key, '')
            result[d] = entry
        else:
            result[d] = None
        d += one_day
    filled = sum(1 for v in result.values() if v is not None)
    print(f'  Calendar built: {filled}/365 days filled (gap = {365-filled})')
    return result

print('Building 2026 calendars...')
adrar_cal   = build_2026_calendar(adrar_times, adrar_hijri)
algiers_cal = build_2026_calendar(algiers_times, algiers_hijri)
djelfa_cal  = build_2026_calendar(djelfa_times, djelfa_hijri)

# ── Step 3: Read offset tables ────────────────────────────────────────────────

def _make_clean_lookup(ar_to_en):
    """Build {clean_arabic(ar): en} so lookups survive alef-variant normalization."""
    return {clean_arabic(ar): en for ar, en in ar_to_en.items()}

_DJELFA_CLEAN  = _make_clean_lookup(DJELFA_AR_TO_EN)
_ALGIERS_CLEAN = _make_clean_lookup(ALGIERS_AR_TO_EN)
_PRAYER_CLEAN  = {clean_arabic(k): v for k, v in PRAYER_AR_MAP.items()}

def read_djelfa_offsets():
    """
    Returns: {city_en: {hijri_month: {prayer: offset_minutes}}}
    """
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'DJEFLFA OFSET.xlsx'), data_only=True)
    ws = wb.active
    headers = [clean_arabic(str(c)) if c else None for c in list(ws.iter_rows(values_only=True))[0]]
    col_to_city = {}
    for i, h in enumerate(headers):
        if i < 2 or not h:
            continue
        city_en = _DJELFA_CLEAN.get(h)
        if city_en:
            col_to_city[i] = city_en

    result = {en: {} for en in DJELFA_AR_TO_EN.values()}
    for row in ws.iter_rows(min_row=2, values_only=True):
        month = normalize_hijri(str(row[0])) if row[0] else None
        prayer_en = _PRAYER_CLEAN.get(clean_arabic(str(row[1]))) if row[1] else None
        if not month or not prayer_en:
            continue
        for i, city_en in col_to_city.items():
            val = row[i]
            if val is None:
                continue
            try:
                offset = int(float(val))
            except (ValueError, TypeError):
                continue
            if month not in result[city_en]:
                result[city_en][month] = {}
            result[city_en][month][prayer_en] = offset
    return result

def read_algiers_offsets():
    """
    Returns: {city_en: {hijri_month: {prayer: offset_minutes}}}
    Long format: (month, prayer, city_ar, offset) per row
    """
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'فروقات_الجزائر.xlsx'), data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        month_ar, prayer_ar, city_ar, val = row[0], row[1], row[2], row[3]
        if not month_ar or not prayer_ar or not city_ar or val is None:
            continue
        month = normalize_hijri(str(month_ar))
        prayer_en = _PRAYER_CLEAN.get(clean_arabic(str(prayer_ar)))
        city_en = _ALGIERS_CLEAN.get(clean_arabic(str(city_ar)))
        if not prayer_en or not city_en:
            continue
        try:
            offset = int(float(val))
        except (ValueError, TypeError):
            continue
        if city_en not in result:
            result[city_en] = {}
        if month not in result[city_en]:
            result[city_en][month] = {}
        result[city_en][month][prayer_en] = offset
    return result

print('Reading offset tables...')
djelfa_offsets  = read_djelfa_offsets()
algiers_offsets = read_algiers_offsets()

# Verify
dj_cities_loaded = [c for c in DJELFA_AR_TO_EN.values() if djelfa_offsets.get(c)]
al_cities_loaded = [c for c in ALGIERS_AR_TO_EN.values() if algiers_offsets.get(c)]
print(f'  Djelfa offsets: {len(dj_cities_loaded)}/26 cities loaded')
print(f'  Algiers offsets: {len(al_cities_loaded)}/25 cities loaded')

# ── Step 4: Build city rows for 2026 ─────────────────────────────────────────

def apply_offset(base_entry, city_en, offsets_dict):
    """Apply monthly offset to base entry. Returns dict of prayer→minutes or None."""
    if base_entry is None:
        return None
    month = base_entry.get('hijri_month', '')
    city_offsets = offsets_dict.get(city_en, {})
    month_offsets = city_offsets.get(month, {})
    result = {}
    for p in PRAYER_COLS:
        base_val = base_entry.get(p)
        if base_val is None:
            result[p] = None
            continue
        off = month_offsets.get(p, 0)
        result[p] = base_val + off
    return result

def cal_to_rows(city_en, cal, offsets_dict=None, is_base=False):
    """
    Convert calendar to list of CSV row dicts.
    If offsets_dict is provided, apply offset to each day.
    Gap days (None entries) are excluded (kept from existing CSV).
    """
    rows = []
    for d, entry in sorted(cal.items()):
        if entry is None:
            continue  # gap: keep current
        if is_base:
            prayers = {p: entry.get(p) for p in PRAYER_COLS}
        else:
            prayers = apply_offset(entry, city_en, offsets_dict)
            if prayers is None:
                continue
        date_str = f'{d.day:02d}/{d.month:02d}/{d.year}'
        row = {'City': city_en, 'Date': date_str}
        for p in PRAYER_COLS:
            row[p.capitalize()] = minutes_to_str(prayers.get(p))
        rows.append(row)
    return rows

print('Building city rows...')

# Base cities
adrar_rows   = cal_to_rows('Adrar',   adrar_cal,   is_base=True)
algiers_rows = cal_to_rows('Algiers', algiers_cal, is_base=True)
djelfa_rows  = cal_to_rows('Djelfa',  djelfa_cal,  is_base=True)

print(f'  Adrar base: {len(adrar_rows)} rows')
print(f'  Algiers base: {len(algiers_rows)} rows')
print(f'  Djelfa base: {len(djelfa_rows)} rows')

# Djelfa surrounding cities
djelfa_city_rows = {}
for city_en in DJELFA_AR_TO_EN.values():
    rows = cal_to_rows(city_en, djelfa_cal, djelfa_offsets)
    djelfa_city_rows[city_en] = rows
    print(f'  Djelfa → {city_en}: {len(rows)} rows')

# Algiers surrounding cities
algiers_city_rows = {}
for city_en in ALGIERS_AR_TO_EN.values():
    rows = cal_to_rows(city_en, algiers_cal, algiers_offsets)
    algiers_city_rows[city_en] = rows
    print(f'  Algiers → {city_en}: {len(rows)} rows')

# ── Step 5: Update CSV ────────────────────────────────────────────────────────

# All cities we are replacing
UPDATED_CITIES = (
    {'Adrar', 'Algiers', 'Djelfa'} |
    set(DJELFA_AR_TO_EN.values()) |
    set(ALGIERS_AR_TO_EN.values())
)
print(f'\nTotal cities to update: {len(UPDATED_CITIES)}')

# Read existing CSV, keep gap rows for updated cities + all rows for other cities
print('Reading existing CSV...')
existing_rows = []  # rows to keep (non-updated cities + gap rows for updated)
gap_dates = {date(2026, 6, d) for d in range(16, 27)}  # Jun 16-26

with open(CSV_PATH, encoding='utf-8', newline='') as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    for row in reader:
        city = row['City']
        if city not in UPDATED_CITIES:
            existing_rows.append(row)
        else:
            # Keep gap rows
            try:
                parts = row['Date'].split('/')
                d = date(int(parts[2]), int(parts[1]), int(parts[0]))
                if d in gap_dates:
                    existing_rows.append(row)
            except Exception:
                pass

print(f'  Kept {len(existing_rows)} rows from existing CSV (non-updated + gaps)')

# Collect all new rows
new_rows = (adrar_rows + algiers_rows + djelfa_rows +
            [r for rows in djelfa_city_rows.values() for r in rows] +
            [r for rows in algiers_city_rows.values() for r in rows])
print(f'  New rows to write: {len(new_rows)}')

# Normalize fieldnames: CSV uses capitalized prayer names
CSV_FIELDS = ['City', 'Date', 'Fajr', 'Sunrise', 'Dhuhr', 'Asr', 'Maghrib', 'Isha']

# Combine and sort by city then date
def sort_key(r):
    city = r['City']
    try:
        parts = r['Date'].split('/')
        d = date(int(parts[2]), int(parts[1]), int(parts[0]))
    except Exception:
        d = date(2026, 1, 1)
    return (city, d)

all_rows = existing_rows + new_rows
all_rows.sort(key=sort_key)

# Write
print('Writing CSV...')
with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
    writer.writeheader()
    for row in all_rows:
        # Normalize keys
        out = {}
        for k in CSV_FIELDS:
            out[k] = row.get(k) or row.get(k.lower()) or ''
        writer.writerow(out)

print(f'Done. Total rows: {len(all_rows)}')

# ── Step 6: Sanity check ──────────────────────────────────────────────────────
print('\n── Sanity check ──')
# Spot check: Djelfa Jan 1, verify dhuhr is in range 12:00-13:30
from collections import defaultdict
city_counts = defaultdict(int)
with open(CSV_PATH, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        city_counts[row['City']] += 1

for city in ['Adrar', 'Algiers', 'Djelfa', 'Ain Temouchent', 'Tlemcen', 'Oran', 'El Tarf']:
    print(f'  {city}: {city_counts.get(city, 0)} rows')

# Check for impossible values (dhuhr should be 11:30-13:30 = 690-810 min)
errors = 0
with open(CSV_PATH, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        city = row['City']
        if city not in UPDATED_CITIES:
            continue
        dhuhr_str = row.get('Dhuhr', '')
        if not dhuhr_str:
            continue
        try:
            parts = dhuhr_str.split(':')
            m = int(parts[0]) * 60 + int(parts[1])
            if m < 690 or m > 810:  # outside 11:30-13:30
                print(f'  WARN: {city} {row["Date"]} dhuhr={dhuhr_str} ({m} min)')
                errors += 1
        except Exception:
            pass

if errors == 0:
    print('  All dhuhr values in valid range (11:30-13:30) ✓')
else:
    print(f'  {errors} suspicious dhuhr values!')

print('\nAll done!')
