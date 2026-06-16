"""
Rebuild Djelfa (base) + its 26 surrounding wilayas from the official 1448 booklet:
  - مواقيت_الصلاة_الجلفة_1448.xlsx   (official Djelfa base, continuous Hijri year
                                      1448: 355 days 2026/06/16..2027/06/05; cols
                                      date|hijri_month|hijri_day|day|fajr|sunrise|
                                      dhuhr|asr|maghrib|isha — Hijri month INLINE
                                      in col[1], prayers row[4..9])
  - فروقات_المدن_الجلفة_1448.xlsx    (official offsets: a 'دليل' sheet + 12 month
                                      sheets named 'محرّم 1448'.. ; each sheet
                                      header المدينة|6 prayers, then 26 city rows.
                                      WARNING on the دليل sheet: numbers were hand-
                                      extracted from small scans — validate.)

city[date] = Djelfa_base[date] + offset[hijri_month_of_date][city][prayer]

Date -> calendar-2026 mapping (single reusable 2026 calendar, no DST in Algeria):
  2026/01/01..06/05 -> xlsx 2027MMDD rows
  2026/06/06..06/15 -> GAP (booklet covers neither year): keep current CSV values
  2026/06/16..12/31 -> xlsx 2026MMDD rows
Maghrib base already carries the official +3; offset applies on top (no extra +3).

Overwrites only Djelfa + the 26 offset cities in assets/csv/algeria_prayer_times_2026.csv.
Dry-run by default (prints validation + diff). Pass --write to persist.
Run:  python tool/rebuild_djelfa_from_1448_xlsx.py [--write]
"""
import sys, io, os, re, csv, unicodedata
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH  = os.path.join(BASE, 'assets', 'csv', 'algeria_prayer_times_2026.csv')
BASE_XLSX = os.path.join(BASE, 'مواقيت_الصلاة_الجلفة_1448.xlsx')
OFF_XLSX  = os.path.join(BASE, 'فروقات_المدن_الجلفة_1448.xlsx')


def clean(s):
    # strip Arabic diacritics (Unicode nonspacing marks) — letters preserved.
    s = ''.join(c for c in (s or '') if unicodedata.category(c) != 'Mn').strip()
    return s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')


PRAYERS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR = {clean(k): v for k, v in {
    'الفجر': 'fajr', 'الشروق': 'sunrise', 'الظهر': 'dhuhr',
    'العصر': 'asr', 'المغرب': 'maghrib', 'العشاء': 'isha',
}.items()}

MONTHS = [clean(m) for m in [
    'محرم', 'صفر', 'ربيع الاول', 'ربيع الثاني', 'جمادى الاولى', 'جمادى الثانية',
    'رجب', 'شعبان', 'رمضان', 'شوال', 'ذو القعدة', 'ذو الحجة',
]]


def month_key(s):
    c = re.sub(r'[0-9٠-٩]+', '', clean(s)).strip()
    return c if c in MONTHS else None


# 26 Djelfa-booklet cities: new-file Arabic label -> CSV English name (all exist).
# (NEW 1448 file uses STANDARD labels — مغنية/سبدو/عين تموشنت/النعامة — unlike the
#  2026-06-10 file's quirky متليلي/سيدو/عين توتة/العلمة.)
DJELFA_AR_TO_EN = {clean(k): v for k, v in {
    'المغير': 'El Mghair', 'أولاد جلال': 'Oulad Djallal', 'مغنية': 'Maghnia',
    'سبدو': 'Sebdou', 'عين تموشنت': 'Ain Temouchent', 'تلمسان': 'Tlemcen',
    'ابن باديس': 'Ben Badis', 'سيدي بلعباس': 'Sidi Bel Abbes', 'النعامة': 'Naama',
    'معسكر': 'Mascara', 'سعيدة': 'Saida', 'البيض': 'El Bayadh', 'تيارت': 'Tiaret',
    'تيسمسيلت': 'Tissemsilt', 'عين وسارة': 'Ain Wasara', 'الأغواط': 'Al-Aghwat',
    'حاسي الرمل': 'Hassi Rmel', 'عين الملح': 'Ain el Melh', 'بوسعادة': 'Bousaada',
    'بسكرة': 'Biskra', 'تقرت': 'Tuggurt', 'باتنة': 'Batna', 'الوادي': 'El Oued',
    'خنشلة': 'Khenchela', 'بئر العاتر': 'Bir al-Ater', 'تبسة': 'Tebessa',
}.items()}


def ptime(s):
    if s is None:
        return None
    m = re.match(r'(\d{1,2}):(\d{2})', str(s).strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def pstr(m):
    return '' if m is None else f'{m // 60:02d}:{m % 60:02d}'


# ── read Djelfa base (continuous 1448 Hijri year, month inline in col[1]) ──────
def read_base():
    ws = openpyxl.load_workbook(BASE_XLSX, data_only=True).active
    times, hijri = {}, {}
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        m = re.match(r'(\d{4})/(\d{2})/(\d{2})', str(a))
        if not m:
            continue
        key = m.group(1) + m.group(2) + m.group(3)
        vals = [ptime(row[4 + i]) for i in range(6)]   # prayers at col 4..9
        if vals[0] is None:
            continue
        times[key] = dict(zip(PRAYERS, vals))
        hijri[key] = month_key(str(row[1]))             # inline Hijri month
    return times, hijri


# ── read offsets (one sheet per Hijri month) ──────────────────────────────────
def read_offsets():
    wb = openpyxl.load_workbook(OFF_XLSX, data_only=True)
    off = {}            # city_en -> month -> prayer -> int
    for shname in wb.sheetnames:
        mk = month_key(shname)
        if not mk:
            continue    # skip the 'دليل' guide sheet
        rows = list(wb[shname].iter_rows(values_only=True))
        pcol = {}       # prayer -> column index, from the header row
        for r in rows:
            if r[0] and clean(str(r[0])) == clean('المدينة'):
                for ci in range(1, len(r)):
                    pk = PRAYER_AR.get(clean(str(r[ci]))) if r[ci] else None
                    if pk:
                        pcol[pk] = ci
                break
        for r in rows:
            if not r[0]:
                continue
            city = DJELFA_AR_TO_EN.get(clean(str(r[0])))
            if city is None:
                continue
            for pk, ci in pcol.items():
                if ci >= len(r) or r[ci] is None:
                    continue
                try:
                    off.setdefault(city, {}).setdefault(mk, {})[pk] = \
                        int(round(float(r[ci])))
                except (ValueError, TypeError):
                    continue
    return off


# ── build calendar-2026 for the Djelfa base ───────────────────────────────────
def build_cal(times, hijri):
    cal = {}
    d = date(2026, 1, 1)
    while d <= date(2026, 12, 31):
        if d >= date(2026, 6, 16):
            k = f'2026{d.month:02d}{d.day:02d}'
        elif d <= date(2026, 6, 5):
            k = f'2027{d.month:02d}{d.day:02d}'
        else:
            cal[d] = None
            d += timedelta(days=1)
            continue
        row = times.get(k)
        cal[d] = ({**row, 'hijri': hijri.get(k, '')} if row else None)
        d += timedelta(days=1)
    return cal


GAP = {date(2026, 6, d) for d in range(6, 16)}          # 06/06..06/15 keep-current
UPD = set(DJELFA_AR_TO_EN.values()) | {'Djelfa'}
LABELS = ['Fajr', 'Sunrise', 'Dhuhr', 'Asr', 'Maghrib', 'Isha']


def main():
    write = '--write' in sys.argv
    print('Reading base + offsets...')
    times, hijri = read_base()
    off = read_offsets()
    print(f'  base days={len(times)} (first={min(times)} last={max(times)})  '
          f'offset cities={len(off)}/26')

    missing = [c for c in DJELFA_AR_TO_EN.values() if c not in off]
    if missing:
        print('  !! MISSING offset cities:', missing)
    for c, mo in off.items():
        bad = [m for m, pp in mo.items() if len(pp) != 6]
        if len(mo) != 12 or bad:
            print(f'  !! {c}: months={len(mo)} incomplete={bad}')

    cal = build_cal(times, hijri)
    filled = sum(1 for v in cal.values() if v)
    print(f'  calendar 2026: {filled}/365 filled, gap={365 - filled} '
          f'({sorted(d.strftime("%d/%m") for d, v in cal.items() if v is None)})')

    city_rows = {}
    for city in DJELFA_AR_TO_EN.values():
        rows = {}
        for d, e in cal.items():
            if e is None:
                continue
            mo = off.get(city, {}).get(e['hijri'], {})
            rows[d] = [e[p] + mo.get(p, 0) for p in PRAYERS]
        city_rows[city] = rows
    djelfa_rows = {d: [e[p] for p in PRAYERS] for d, e in cal.items() if e}

    # ── VALIDATION ────────────────────────────────────────────────────────────
    print('\n── validation ──')
    issues = 0
    all_rows_for_check = list(city_rows.items()) + [('Djelfa', djelfa_rows)]
    for city, rows in all_rows_for_check:
        for d, t in rows.items():
            if any(x is None for x in t):
                continue
            if not (t[0] < t[1] < t[2] < t[3] < t[4] < t[5]):
                print(f'  ORDER  {city} {d}: {[pstr(x) for x in t]}')
                issues += 1
    TH = [12, 12, 8, 12, 10, 14]
    for city, rows in all_rows_for_check:
        ds = sorted(rows.keys())
        for a, b in zip(ds, ds[1:]):
            if (b - a).days != 1:
                continue
            for i in range(6):
                j = rows[b][i] - rows[a][i]
                if abs(j) > TH[i]:
                    print(f'  JUMP   {city} {LABELS[i]} {a}->{b}: '
                          f'{pstr(rows[a][i])}->{pstr(rows[b][i])} ({j:+d})')
                    issues += 1
    print(f'  structural issues: {issues}')

    # diff vs current CSV (max delta flags hand-extraction outliers)
    cur = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r[0] in UPD and '/' in r[1]:
                dd, mm, yy = r[1].split('/')
                cur[(r[0], date(int(yy), int(mm), int(dd)))] = [ptime(x) for x in r[2:8]]
    changed = maxch = 0
    per_city_max = {}
    samples = []
    for city, rows in all_rows_for_check:
        for d, t in rows.items():
            c = cur.get((city, d))
            if not c:
                continue
            for i in range(6):
                if t[i] is not None and c[i] is not None and t[i] != c[i]:
                    dl = abs(t[i] - c[i])
                    changed += 1
                    maxch = max(maxch, dl)
                    per_city_max[city] = max(per_city_max.get(city, 0), dl)
                    if dl >= 10 and len(samples) < 20:
                        samples.append(f'  {city} {d} {LABELS[i]}: '
                                       f'{pstr(c[i])}->{pstr(t[i])} ({t[i] - c[i]:+d})')
    print(f'\n  cells changed vs current CSV: {changed}  (max delta {maxch} min)')
    print('  per-city max delta (min):')
    for city in sorted(per_city_max, key=lambda c: -per_city_max[c]):
        print(f'    {city:18} {per_city_max[city]}')
    if samples:
        print('  large changes (>=10 min) to eyeball:')
        for sline in samples:
            print(sline)

    print('\n  gap seam (Djelfa, new vs kept-current):')
    for d in (date(2026, 6, 5), date(2026, 6, 6), date(2026, 6, 15), date(2026, 6, 16)):
        src = 'kept' if d in GAP else 'new '
        t = djelfa_rows.get(d) or cur.get(('Djelfa', d))
        print(f'    {d} [{src}] {[pstr(x) for x in t]}')

    if issues:
        print('\n!! structural issues present — NOT writing CSV. Inspect first.')
        return
    if not write:
        print('\n[dry-run] add --write to overwrite the CSV.')
        return

    # ── WRITE CSV ─────────────────────────────────────────────────────────────
    keep = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        rd = csv.reader(f)
        header = next(rd)
        for r in rd:
            if r[0] not in UPD:
                keep.append(r)
                continue
            dd, mm, yy = r[1].split('/')
            if date(int(yy), int(mm), int(dd)) in GAP:
                keep.append(r)

    new = []

    def emit(city, rows):
        for d, t in sorted(rows.items()):
            new.append([city, f'{d.day:02d}/{d.month:02d}/{d.year}']
                       + [pstr(x) for x in t])

    emit('Djelfa', djelfa_rows)
    for city, rows in city_rows.items():
        emit(city, rows)

    from collections import Counter
    cnt = Counter(r[0] for r in keep + new if r[0] in UPD)
    bad = {c: n for c, n in cnt.items() if n != 365}
    if bad:
        print('\n!! row-count mismatch (expected 365):', bad, '— NOT writing.')
        return

    all_rows = keep + new

    def sk(r):
        dd, mm, yy = r[1].split('/')
        return (r[0], date(int(yy), int(mm), int(dd)))

    all_rows.sort(key=sk)
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        wr = csv.writer(f)
        wr.writerow(header)
        wr.writerows(all_rows)
    print(f'\nWrote CSV: kept={len(keep)} new={len(new)} total={len(all_rows)}')


if __name__ == '__main__':
    main()
