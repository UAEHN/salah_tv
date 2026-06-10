"""
Rebuild the 25 Algiers-booklet surrounding cities from:
  - مواقيت_الصلاة_ولاية_الجزائر_العاصمة_1447.xlsx  (official Algiers base, 354 days)
  - فروقات_المواقيت_الجزائر_أرقام.xlsx              (official offset table, 12mo x 6 x 25)

city[date] = Algiers_base[date] + offset[hijri_month_of_date][prayer]

Date->calendar-2026 mapping (per project pref):
  2026-01-01..06-15 -> xlsx 2026MMDD rows
  2026-06-16..06-26 -> GAP: keep current CSV values (do NOT interpolate)
  2026-06-27..12-31 -> xlsx 2025MMDD rows
Maghrib base already carries the official +3; offset applies on top (no extra +3).

Overwrites only Algiers + the 25 cities in assets/csv/algeria_prayer_times_2026.csv.
Run:  python tool/rebuild_algiers_from_new_offsets.py
"""
import sys, io, os, re, csv
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH  = os.path.join(BASE, 'assets', 'csv', 'algeria_prayer_times_2026.csv')
BASE_XLSX = os.path.join(BASE, 'مواقيت_الصلاة_ولاية_الجزائر_العاصمة_1447.xlsx')
OFF_XLSX  = os.path.join(BASE, 'فروقات_المواقيت_الجزائر_أرقام.xlsx')

import unicodedata
def clean(s):
    # strip Arabic diacritics (Unicode nonspacing marks) — letters are preserved
    s = ''.join(c for c in (s or '') if unicodedata.category(c) != 'Mn').strip()
    return s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')

PRAYERS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR = {'الفجر':'fajr','الشروق':'sunrise','الظهر':'dhuhr',
             'العصر':'asr','المغرب':'maghrib','العشاء':'isha'}
PRAYER_AR = {clean(k):v for k,v in PRAYER_AR.items()}

# 25 Algiers-booklet cities: Arabic offset label -> CSV English name
ALGIERS_AR_TO_EN = {
    'وهران':'Oran','مستغانم':'Mostaganem','غليزان':'Relizane','الشلف':'Chlef',
    'عين الدفلى':'Ain Defla','تيبازة':'Tipaza','المدية':'Medea','البليدة':'Blida',
    'بومرداس':'Boumerdes','البويرة':'Bouira','دلس':'Dellys','تيزي وزو':'Tizi Ouzou',
    'المسيلة':'Msila','برج بوعريريج':'Bordj Bou Arreridj','باتنة':'Batna',
    'سطيف':'Setif','جيجل':'Jijel','ميلة':'Milah','قسنطينة':'Constantine',
    'سكيكدة':'Skikda','أم البواقي':'Um al-Bouaki','قالمة':'Qalma','عنابة':'Annaba',
    'سوق أهراس':'Souk Ahras','الطارف':'El Tarf',
}
AR_CLEAN = {clean(k):v for k,v in ALGIERS_AR_TO_EN.items()}

def ptime(s):
    if s is None: return None
    m = re.match(r'(\d{1,2}):(\d{2})', str(s).strip())
    return int(m.group(1))*60+int(m.group(2)) if m else None
def pstr(m):
    return '' if m is None else f'{m//60:02d}:{m%60:02d}'

# ── read Algiers base ─────────────────────────────────────────────────────────
def read_base():
    wb = openpyxl.load_workbook(BASE_XLSX, data_only=True); ws = wb.active
    times, hijri = {}, {}
    cur_month = None
    for row in ws.iter_rows(values_only=True):
        dv = row[1]
        if dv is None or '/' not in str(dv):
            if row[0] and 'محرم' in clean(str(row[0])) or (row[0] and clean(str(row[0])) in
               {'صفر','رجب','رمضان','شوال'}):
                cur_month = clean(str(row[0])).replace('شهر ','')
            continue
        m = re.match(r'(\d{4})/(\d{2})/(\d{2})', str(dv))
        if not m: continue
        key = m.group(1)+m.group(2)+m.group(3)
        vals = [ptime(row[3+i]) for i in range(6)]
        if vals[0] is None: continue
        times[key] = dict(zip(PRAYERS, vals))
        if row[0] and str(row[0]).strip():
            cur_month = clean(str(row[0])).replace('شهر ','')
        hijri[key] = cur_month
    return times, hijri

# ── read offsets (long sheet) ─────────────────────────────────────────────────
def read_offsets():
    wb = openpyxl.load_workbook(OFF_XLSX, data_only=True)
    ws = wb['كل الفروقات']
    off = {}  # city_en -> month -> prayer -> int
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        month = clean(str(row[0])).replace('شهر ','') if row[0] else None
        prayer = PRAYER_AR.get(clean(str(row[1]))) if row[1] else None
        city = AR_CLEAN.get(clean(str(row[2]))) if row[2] else None
        if not (month and prayer and city) or row[3] is None: continue
        try: val = int(float(row[3]))
        except (ValueError, TypeError): continue
        off.setdefault(city, {}).setdefault(month, {})[prayer] = val
    return off

# ── build calendar 2026 for base ──────────────────────────────────────────────
def build_cal(times, hijri):
    cal = {}; d = date(2026,1,1)
    while d <= date(2026,12,31):
        if d <= date(2026,6,15): k = f'2026{d.month:02d}{d.day:02d}'
        elif d <= date(2026,6,26): cal[d] = None; d += timedelta(days=1); continue
        else: k = f'2025{d.month:02d}{d.day:02d}'
        row = times.get(k)
        if row:
            e = dict(row); e['hijri'] = hijri.get(k, '')
            cal[d] = e
        else:
            cal[d] = None
        d += timedelta(days=1)
    return cal

def main():
    print('Reading base + offsets...')
    times, hijri = read_base(); off = read_offsets()
    print(f'  base days={len(times)}  offset cities={len(off)}/25')
    missing = [c for c in ALGIERS_AR_TO_EN.values() if c not in off]
    if missing: print('  !! MISSING offset cities:', missing)
    # verify each loaded city has 12 months x 6 prayers
    for c, mo in off.items():
        bad = [m for m,pp in mo.items() if len(pp)!=6]
        if len(mo)!=12 or bad:
            print(f'  !! {c}: months={len(mo)} incomplete={bad}')

    cal = build_cal(times, hijri)
    filled = sum(1 for v in cal.values() if v)
    print(f'  calendar 2026: {filled}/365 filled, gap={365-filled}')

    # build city rows
    city_rows = {}   # city_en -> {date: [6 mins]}
    for city in ALGIERS_AR_TO_EN.values():
        rows = {}
        for d, e in cal.items():
            if e is None: continue
            mo = off[city].get(e['hijri'], {})
            rows[d] = [e[p] + mo.get(p, 0) for p in PRAYERS]
        city_rows[city] = rows
    # Algiers base itself (unchanged, but rewrite cleanly)
    alg_rows = {d:[e[p] for p in PRAYERS] for d,e in cal.items() if e}

    # ── VALIDATION ────────────────────────────────────────────────────────────
    print('\n── validation ──')
    issues = 0
    labels = ['Fajr','Sunrise','Dhuhr','Asr','Maghrib','Isha']
    # 1) monotonic ordering each day
    for city, rows in list(city_rows.items()) + [('Algiers', alg_rows)]:
        for d, t in rows.items():
            if any(x is None for x in t): continue
            if not (t[0]<t[1]<t[2]<t[3]<t[4]<t[5]):
                print(f'  ORDER  {city} {d}: {[pstr(x) for x in t]}'); issues += 1
    # 2) day-to-day smoothness (incl. month boundaries) per prayer
    TH = [12,12,8,12,10,14]
    for city, rows in city_rows.items():
        ds = sorted(rows.keys())
        for a, b in zip(ds, ds[1:]):
            if (b-a).days != 1: continue   # skip across the gap
            for i in range(6):
                j = rows[b][i]-rows[a][i]
                if abs(j) > TH[i]:
                    print(f'  JUMP   {city} {labels[i]} {a}->{b}: '
                          f'{pstr(rows[a][i])}->{pstr(rows[b][i])} ({j:+d})'); issues += 1
    print(f'  structural issues: {issues}')

    # 3) diff vs current CSV (effective offset change report)
    cur = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r[0] in ALGIERS_AR_TO_EN.values():
                dd,mm,yy = r[1].split('/')
                cur[(r[0], date(int(yy),int(mm),int(dd)))] = [ptime(x) for x in r[2:8]]
    changed = 0; maxch = 0; samples = []
    for city, rows in city_rows.items():
        for d, t in rows.items():
            c = cur.get((city, d))
            if not c: continue
            for i in range(6):
                if t[i] is not None and c[i] is not None and t[i]!=c[i]:
                    changed += 1; maxch = max(maxch, abs(t[i]-c[i]))
                    if len(samples) < 12:
                        samples.append(f'  {city} {d} {labels[i]}: '
                                       f'{pstr(c[i])}->{pstr(t[i])} ({t[i]-c[i]:+d})')
    print(f'\n  cells changed vs current CSV: {changed}  (max delta {maxch} min)')
    for s in samples: print(s)

    if issues:
        print('\n!! structural issues present — NOT writing CSV. Inspect first.')
        return

    # ── WRITE CSV ─────────────────────────────────────────────────────────────
    UPD = set(ALGIERS_AR_TO_EN.values()) | {'Algiers'}
    gap = {date(2026,6,d) for d in range(16,27)}
    keep = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        rd = csv.reader(f); header = next(rd)
        for r in rd:
            if r[0] not in UPD: keep.append(r); continue
            dd,mm,yy = r[1].split('/')
            if date(int(yy),int(mm),int(dd)) in gap: keep.append(r)  # keep-current gap
    new = []
    def emit(city, rows):
        for d, t in sorted(rows.items()):
            new.append([city, f'{d.day:02d}/{d.month:02d}/{d.year}'] + [pstr(x) for x in t])
    emit('Algiers', alg_rows)
    for city, rows in city_rows.items(): emit(city, rows)

    all_rows = keep + new
    def sk(r):
        dd,mm,yy = r[1].split('/'); return (r[0], date(int(yy),int(mm),int(dd)))
    all_rows.sort(key=sk)
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        wr = csv.writer(f); wr.writerow(header); wr.writerows(all_rows)
    print(f'\nWrote CSV: kept={len(keep)} new={len(new)} total={len(all_rows)}')

if __name__ == '__main__':
    main()
