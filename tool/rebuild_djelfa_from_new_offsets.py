"""
Rebuild the Djelfa base + its 26 surrounding cities from:
  - مواقيت_الصلاة_ولاية_الجلفة_1447.xlsx  (official Djelfa base, 354 days)
  - furoqat_djelfa_all_months.xlsx         (official offset table, 12mo x 6 x 26)
    long sheet "كل الفروقات" cols: [month, gregorian_period, prayer, city, offset]

city[date] = Djelfa_base[date] + offset[hijri_month_of_date][prayer]

City identities were verified by longitude calc (dhuhr_off ≈ (Djelfa_lng - city_lng)*4,
all 25 with coords matched within ±1.4 min). The file uses NON-STANDARD labels for
4 cities — values, not labels, identify them:
  متليلي = Maghnia, سيدو = Sebdou, العلمة = Naama, عين توتة = Ain Temouchent.

Date->2026 mapping (per project pref):
  01-01..06-15 -> xlsx 2026 rows ; 06-16..06-26 -> GAP ; 06-27..12-31 -> xlsx 2025 rows
Gap days are filled from current-CSV Djelfa base + ذو الحجة offset (avoids a mid-year
hole; the 354-day xlsx does not cover 16-26 June).
Maghrib base already carries +3; offset applies on top (no extra +3).

Overwrites only Djelfa + the 26 cities in assets/csv/algeria_prayer_times_2026.csv.
Run:  python tool/rebuild_djelfa_from_new_offsets.py
"""
import sys, io, os, re, csv, unicodedata
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH  = os.path.join(BASE, 'assets', 'csv', 'algeria_prayer_times_2026.csv')
BASE_XLSX = os.path.join(BASE, 'مواقيت_الصلاة_ولاية_الجلفة_1447.xlsx')
OFF_XLSX  = os.path.join(BASE, 'furoqat_djelfa_all_months.xlsx')

def clean(s):
    s = ''.join(c for c in (s or '') if unicodedata.category(c) != 'Mn').strip()
    return s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')

PRAYERS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR = {clean(k):v for k,v in {'الفجر':'fajr','الشروق':'sunrise','الظهر':'dhuhr',
             'العصر':'asr','المغرب':'maghrib','العشاء':'isha'}.items()}

# NEW-file Arabic label -> CSV English name (verified by longitude calc)
DJELFA_AR_TO_EN = {
    'المغير':'El Mghair','أولاد جلال':'Oulad Djallal','متليلي':'Maghnia','سيدو':'Sebdou',
    'عين توتة':'Ain Temouchent','تلمسان':'Tlemcen','ابن باديس':'Ben Badis',
    'سيدي بلعباس':'Sidi Bel Abbes','العلمة':'Naama','معسكر':'Mascara','سعيدة':'Saida',
    'البيض':'El Bayadh','تيارت':'Tiaret','تيسمسيلت':'Tissemsilt','عين وسارة':'Ain Wasara',
    'الأغواط':'Al-Aghwat','حاسي الرمل':'Hassi Rmel','عين الملح':'Ain el Melh',
    'بوسعادة':'Bousaada','بسكرة':'Biskra','تقرت':'Tuggurt','باتنة':'Batna',
    'الوادي':'El Oued','خنشلة':'Khenchela','بئر العاتر':'Bir al-Ater','تبسة':'Tebessa',
}
AR_CLEAN = {clean(k):v for k,v in DJELFA_AR_TO_EN.items()}
GAP_MONTH = clean('ذو الحجة')  # hijri month covering the 16-26 June gap

def ptime(s):
    if s is None: return None
    m = re.match(r'(\d{1,2}):(\d{2})', str(s).strip())
    return int(m.group(1))*60+int(m.group(2)) if m else None
def pstr(m):
    return '' if m is None else f'{m//60:02d}:{m%60:02d}'

def read_base():
    wb = openpyxl.load_workbook(BASE_XLSX, data_only=True); ws = wb.active
    times, hijri = {}, {}; cur = None
    for row in ws.iter_rows(values_only=True):
        dv = row[1]
        if dv is None or not re.match(r'20\d\d/\d\d/\d\d', str(dv)):
            if row[0] and str(row[0]).strip():
                t = clean(str(row[0])).replace('شهر ','')
                if t in {'محرم','صفر','ربيع الاول','ربيع الثاني','جمادى الاولى',
                         'جمادى الثانية','رجب','شعبان','رمضان','شوال','ذو القعدة','ذو الحجة'}:
                    cur = t
            continue
        m = re.match(r'(\d{4})/(\d{2})/(\d{2})', str(dv))
        key = m.group(1)+m.group(2)+m.group(3)
        vals = [ptime(row[3+i]) for i in range(6)]
        if vals[0] is None: continue
        times[key] = dict(zip(PRAYERS, vals))
        if row[0] and str(row[0]).strip():
            cur = clean(str(row[0])).replace('شهر ','')
        hijri[key] = cur
    return times, hijri

def read_offsets():
    wb = openpyxl.load_workbook(OFF_XLSX, data_only=True)
    ws = wb['كل الفروقات']   # cols: month, period, prayer, city, offset
    off = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        month = clean(str(row[0])).replace('شهر ','') if row[0] else None
        prayer = PRAYER_AR.get(clean(str(row[2]))) if row[2] else None
        city = AR_CLEAN.get(clean(str(row[3]))) if row[3] else None
        if not (month and prayer and city) or row[4] is None: continue
        try: val = int(float(row[4]))
        except (ValueError, TypeError): continue
        off.setdefault(city, {}).setdefault(month, {})[prayer] = val
    return off

def build_cal(times, hijri):
    cal = {}; d = date(2026,1,1)
    while d <= date(2026,12,31):
        if d <= date(2026,6,15): k = f'2026{d.month:02d}{d.day:02d}'
        elif d <= date(2026,6,26): cal[d] = None; d += timedelta(days=1); continue
        else: k = f'2025{d.month:02d}{d.day:02d}'
        row = times.get(k)
        if row:
            e = dict(row); e['hijri'] = hijri.get(k, ''); cal[d] = e
        else: cal[d] = None
        d += timedelta(days=1)
    return cal

def read_current_gap_base():
    """Djelfa base values for the 16-26 June gap, from the existing CSV."""
    gap = {date(2026,6,dd) for dd in range(16,27)}
    out = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r and r[0] == 'Djelfa':
                dd,mm,yy = r[1].split('/'); d = date(int(yy),int(mm),int(dd))
                if d in gap: out[d] = dict(zip(PRAYERS, [ptime(x) for x in r[2:8]]))
    return out

def main():
    print('Reading Djelfa base + offsets...')
    times, hijri = read_base(); off = read_offsets()
    print(f'  base days={len(times)}  offset cities={len(off)}/26')
    miss = [c for c in DJELFA_AR_TO_EN.values() if c not in off]
    if miss: print('  !! MISSING:', miss)
    for c, mo in off.items():
        bad = [m for m,pp in mo.items() if len(pp)!=6]
        if len(mo)!=12 or bad: print(f'  !! {c}: months={len(mo)} bad={bad}')

    cal = build_cal(times, hijri)
    gap_base = read_current_gap_base()
    print(f'  calendar filled={sum(1 for v in cal.values() if v)}/365  gap_base_days={len(gap_base)}')

    # build all city rows (full 365 incl gap)
    def city_series(city):
        rows = {}
        for d, e in cal.items():
            if e is None: continue
            mo = off[city].get(e['hijri'], {})
            rows[d] = [e[p] + mo.get(p, 0) for p in PRAYERS]
        gmo = off[city].get(GAP_MONTH, {})
        for d, base in gap_base.items():
            rows[d] = [base[p] + gmo.get(p, 0) for p in PRAYERS]
        return rows
    city_rows = {c: city_series(c) for c in DJELFA_AR_TO_EN.values()}
    # Djelfa base itself: xlsx non-gap + current gap
    djelfa_rows = {d:[e[p] for p in PRAYERS] for d,e in cal.items() if e}
    for d, base in gap_base.items():
        djelfa_rows[d] = [base[p] for p in PRAYERS]

    # ── validation ────────────────────────────────────────────────────────────
    print('\n── validation ──')
    issues = 0; labels = ['Fajr','Sun','Dhuhr','Asr','Maghrib','Isha']
    TH = [13,13,9,13,11,15]
    for city, rows in list(city_rows.items()) + [('Djelfa', djelfa_rows)]:
        if len(rows) != 365:
            print(f'  COUNT {city}: {len(rows)} rows'); issues += 1
        for d, t in rows.items():
            if any(x is None for x in t) or not (t[0]<t[1]<t[2]<t[3]<t[4]<t[5]):
                print(f'  ORDER {city} {d}: {[pstr(x) for x in t]}'); issues += 1
        ds = sorted(rows)
        for a, b in zip(ds, ds[1:]):
            if (b-a).days != 1: continue
            for i in range(6):
                j = rows[b][i]-rows[a][i]
                if abs(j) > TH[i]:
                    print(f'  JUMP {city} {labels[i]} {a}->{b}: '
                          f'{pstr(rows[a][i])}->{pstr(rows[b][i])} ({j:+d})'); issues += 1
    print(f'  structural issues: {issues}')

    # diff vs current
    cur = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r and r[0] in (set(DJELFA_AR_TO_EN.values()) | {'Djelfa'}):
                dd,mm,yy = r[1].split('/')
                cur[(r[0], date(int(yy),int(mm),int(dd)))] = [ptime(x) for x in r[2:8]]
    changed = maxch = 0
    allrows = dict(city_rows); allrows['Djelfa'] = djelfa_rows
    for city, rows in allrows.items():
        for d, t in rows.items():
            c = cur.get((city, d))
            if not c: continue
            for i in range(6):
                if t[i] is not None and c[i] is not None and t[i]!=c[i]:
                    changed += 1; maxch = max(maxch, abs(t[i]-c[i]))
    print(f'  cells changed vs current CSV: {changed} (max delta {maxch} min)')

    if issues:
        print('\n!! issues present — NOT writing.'); return

    # ── write ─────────────────────────────────────────────────────────────────
    UPD = set(DJELFA_AR_TO_EN.values()) | {'Djelfa'}
    keep = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        rd = csv.reader(f); header = next(rd)
        for r in rd:
            if r[0] not in UPD: keep.append(r)
    new = []
    for city, rows in allrows.items():
        for d, t in sorted(rows.items()):
            new.append([city, f'{d.day:02d}/{d.month:02d}/{d.year}'] + [pstr(x) for x in t])
    out = keep + new
    out.sort(key=lambda r: (r[0], date(*map(int, r[1].split('/')[::-1]))))
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        wr = csv.writer(f); wr.writerow(header); wr.writerows(out)
    print(f'\nWrote CSV: kept={len(keep)} new={len(new)} total={len(out)}')

if __name__ == '__main__':
    main()
