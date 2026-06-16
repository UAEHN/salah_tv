"""
Rebuild Adrar (base) + its 16 surrounding desert cities from the official 1448
booklet:
  - مواقيت_صلاة_أدرار_1448.xlsx   (official Adrar base, continuous Hijri year 1448:
                                   355 days 2026-06-16..2027-06-05; DASH dates; cols
                                   date|day|hijri_month|hijri_day|fajr|sunrise|dhuhr|
                                   asr|maghrib|isha — month+day INLINE, prayers [4..9])
  - فروق_توقيت_مدن_أدرار_1448.xlsx (official offsets: 'فروق أدرار' sheet + 'تعليمات'.
                                   Each Hijri month has TWO PERIODS (days 1–15, 16–end)
                                   because the offset drifts within the month. Columns:
                                   الشهر|الفترة|الصلاة|16 cities. Dhuhr row is constant
                                   all year (longitude only). Numbers were VISUALLY
                                   extracted from photos — structural validation matters.)

city[date] = Adrar_base[date] + offset[hijri_month][period][city][prayer]
  period = A (hijri day 1–15) | B (hijri day 16–end)

Date -> calendar-2026 mapping (single reusable 2026 calendar, no DST in Algeria):
  2026/01/01..06/05 -> xlsx 2027MMDD rows
  2026/06/06..06/15 -> GAP (booklet covers neither year). Filled = current Adrar
                       base for those days + the ذو الحجة period-B offset (the
                       seasonally-adjacent block, last in the file at 2027-06-05),
                       so BOTH seams stay smooth instead of stitching calc onto
                       official. Adrar city itself keeps its current gap values.
  2026/06/16..12/31 -> xlsx 2026MMDD rows
Maghrib base already carries the official +3; offset applies on top (no extra +3).

Overwrites only Adrar + the 16 offset cities in assets/csv/algeria_prayer_times_2026.csv.
Dry-run by default (prints validation + diff). Pass --write to persist.
Run:  python tool/rebuild_adrar_from_1448_xlsx.py [--write]
"""
import sys, io, os, re, csv, unicodedata
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH  = os.path.join(BASE, 'assets', 'csv', 'algeria_prayer_times_2026.csv')
BASE_XLSX = os.path.join(BASE, 'مواقيت_صلاة_أدرار_1448.xlsx')
OFF_XLSX  = os.path.join(BASE, 'فروق_توقيت_مدن_أدرار_1448.xlsx')


def clean(s):
    s = ''.join(c for c in (s or '') if unicodedata.category(c) != 'Mn').strip()
    return s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')


def norm(s):
    # city-label key: drop spaces, dashes (any unicode), dots, tatweel — so
    # عين-أمناس / عين أمناس / عينأمناس all collapse to one robust key.
    return re.sub(r'[\s\.\-‐‑‒–—―−ـ]', '', clean(s))


PRAYERS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR = {clean(k): v for k, v in {
    'الفجر': 'fajr', 'الشروق': 'sunrise', 'الظهر': 'dhuhr',
    'العصر': 'asr', 'المغرب': 'maghrib', 'العشاء': 'isha',
}.items()}

MONTHS = [clean(m) for m in [
    'محرم', 'صفر', 'ربيع الاول', 'ربيع الثاني', 'جمادى الاولى', 'جمادى الثانية',
    'رجب', 'شعبان', 'رمضان', 'شوال', 'ذو القعدة', 'ذو الحجة',
]]
DHU_HIJJA = clean('ذو الحجة')


def month_key(s):
    c = re.sub(r'[0-9٠-٩]+', '', clean(s)).strip()
    return c if c in MONTHS else None


# 16 Adrar-booklet desert cities: file label -> CSV English name (all exist).
# NOTE: this file's إليزي/جانت COLUMNS ARE SWAPPED (verified two ways: the Dhuhr
# offset is pure longitude, and the إليزي column's full vector matches current-calc
# Djanet, and vice-versa). So the إليزي column actually holds Djanet's offsets and
# the جانت column holds Illizi's — mapped accordingly here. (See the §"identify by
# values not label" gotcha in memory.)
ADRAR_NORM_TO_EN = {norm(k): v for k, v in {
    'عين-أمناس': 'In Amenas', 'إليزي': 'Djanet', 'جانت': 'Illizi',
    'إن-قزام': 'In Guezzam', 'تمنراست': 'Tamanrasset', 'ورقلة': 'Ouargla',
    'غرداية': 'Ghardaia', 'المنيعة': 'El Menia', 'إن-صالح': 'In Salah',
    'بني-عباس': 'Beni Abbes', 'ب.ب.مختار': 'Bordj Badji Mokhtar', 'رقان': 'Reggane',
    'تيميمون': 'Timimoun', 'بشار': 'Bechar', 'تندوف': 'Tindouf',
    'بني-ونيف': 'Beni Ounif',
}.items()}


def ptime(s):
    if s is None:
        return None
    m = re.match(r'(\d{1,2}):(\d{2})', str(s).strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def pstr(m):
    return '' if m is None else f'{m // 60:02d}:{m % 60:02d}'


# ── read Adrar base (dash dates, month+day inline) ────────────────────────────
def read_base():
    ws = openpyxl.load_workbook(BASE_XLSX, data_only=True).active
    times, hijri, hday = {}, {}, {}
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        m = re.match(r'(\d{4})[-/](\d{2})[-/](\d{2})', str(a))
        if not m:
            continue
        key = m.group(1) + m.group(2) + m.group(3)
        vals = [ptime(row[4 + i]) for i in range(6)]   # prayers at col 4..9
        if vals[0] is None:
            continue
        times[key] = dict(zip(PRAYERS, vals))
        hijri[key] = month_key(str(row[2]))
        try:
            hday[key] = int(row[3])
        except (ValueError, TypeError):
            hday[key] = None
    return times, hijri, hday


# ── read offsets (single sheet, month × period × prayer × 16 cities) ───────────
def read_offsets():
    ws = openpyxl.load_workbook(OFF_XLSX, data_only=True)['فروق أدرار']
    rows = list(ws.iter_rows(values_only=True))
    col_city = {}
    for r in rows:
        if r[2] is not None and clean(str(r[2])) == clean('الصلاة'):
            for ci in range(3, len(r)):
                col_city[ci] = ADRAR_NORM_TO_EN.get(norm(str(r[ci]))) if r[ci] else None
            break
    off = {}            # city -> month -> period('A'/'B') -> prayer -> int
    for r in rows:
        if r[0] is None or r[1] is None or r[2] is None:
            continue
        mk = month_key(str(r[0]))
        pk = PRAYER_AR.get(clean(str(r[2])))
        pm = re.match(r'\s*(\d+)', str(r[1]))
        if not (mk and pk and pm):
            continue
        period = 'A' if int(pm.group(1)) == 1 else 'B'
        for ci, city in col_city.items():
            if city is None or ci >= len(r) or r[ci] is None:
                continue
            try:
                off.setdefault(city, {}).setdefault(mk, {}) \
                   .setdefault(period, {})[pk] = int(round(float(r[ci])))
            except (ValueError, TypeError):
                continue
    return off


def off_for(off, city, month, period):
    return off.get(city, {}).get(month, {}).get(period, {})


# ── build calendar-2026 for the Adrar base ────────────────────────────────────
def build_cal(times, hijri, hday):
    cal = {}
    d = date(2026, 1, 1)
    while d <= date(2026, 12, 31):
        if d >= date(2026, 6, 16):
            k = f'2026{d.month:02d}{d.day:02d}'
        elif d <= date(2026, 6, 5):
            k = f'2027{d.month:02d}{d.day:02d}'
        else:
            cal[d] = None
            d += timedelta(days=1)
            continue
        row = times.get(k)
        cal[d] = ({**row, 'hijri': hijri.get(k), 'hday': hday.get(k)} if row else None)
        d += timedelta(days=1)
    return cal


GAP = sorted(date(2026, 6, d) for d in range(6, 16))    # 06/06..06/15
UPD = set(ADRAR_NORM_TO_EN.values()) | {'Adrar'}
LABELS = ['Fajr', 'Sunrise', 'Dhuhr', 'Asr', 'Maghrib', 'Isha']


def main():
    write = '--write' in sys.argv
    print('Reading base + offsets...')
    times, hijri, hday = read_base()
    off = read_offsets()
    print(f'  base days={len(times)} (first={min(times)} last={max(times)})  '
          f'offset cities={len(off)}/16')

    missing = [c for c in ADRAR_NORM_TO_EN.values() if c not in off]
    if missing:
        print('  !! MISSING offset cities:', missing)
    for c, mo in off.items():
        bad = [m for m, pp in mo.items() if set(pp) != {'A', 'B'}
               or any(len(pp[per]) != 6 for per in pp)]
        if len(mo) != 12 or bad:
            print(f'  !! {c}: months={len(mo)} bad={bad}')

    cal = build_cal(times, hijri, hday)
    filled = sum(1 for v in cal.values() if v)
    print(f'  calendar 2026: {filled}/365 filled, gap={365 - filled}')

    # current CSV (need Adrar gap base + diff report)
    cur = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r[0] in UPD and '/' in r[1]:
                dd, mm, yy = r[1].split('/')
                cur[(r[0], date(int(yy), int(mm), int(dd)))] = [ptime(x) for x in r[2:8]]

    # surrounding cities: 355 covered days + 10 gap days (Adrar-base + ذو الحجة-B)
    city_rows = {}
    for city in ADRAR_NORM_TO_EN.values():
        rows = {}
        for d, e in cal.items():
            if e is None:
                continue
            period = 'A' if (e['hday'] or 99) <= 15 else 'B'
            mo = off_for(off, city, e['hijri'], period)
            rows[d] = [e[p] + mo.get(p, 0) for p in PRAYERS]
        gmo = off_for(off, city, DHU_HIJJA, 'B')
        for d in GAP:
            base6 = cur.get(('Adrar', d))
            if base6 and all(x is not None for x in base6):
                rows[d] = [base6[i] + gmo.get(PRAYERS[i], 0) for i in range(6)]
        city_rows[city] = rows

    # Adrar city: 355 official + 10 kept-current gap
    adrar_rows = {d: [e[p] for p in PRAYERS] for d, e in cal.items() if e}
    for d in GAP:
        base6 = cur.get(('Adrar', d))
        if base6:
            adrar_rows[d] = list(base6)

    # ── VALIDATION (full 365-day contiguous → seams ARE checked) ───────────────
    print('\n── validation ──')
    issues = 0
    all_rows = list(city_rows.items()) + [('Adrar', adrar_rows)]
    for city, rows in all_rows:
        for d, t in rows.items():
            if any(x is None for x in t):
                continue
            if not (t[0] < t[1] < t[2] < t[3] < t[4] < t[5]):
                print(f'  ORDER  {city} {d}: {[pstr(x) for x in t]}')
                issues += 1
    TH = [12, 12, 8, 14, 12, 16]   # desert cities drift a touch more at seams
    for city, rows in all_rows:
        ds = sorted(rows.keys())
        for a, b in zip(ds, ds[1:]):
            if (b - a).days != 1:
                continue
            for i in range(6):
                j = rows[b][i] - rows[a][i]
                if abs(j) > TH[i]:
                    print(f'  JUMP   {city} {LABELS[i]} {a}->{b}: '
                          f'{pstr(rows[a][i])}->{pstr(rows[b][i])} ({j:+d})')
                    issues += 1
    print(f'  structural issues: {issues}')

    # diff vs current (EXPECTED large: switching calc -> official for far cities)
    per_city_max = {}
    changed = 0
    for city, rows in all_rows:
        for d, t in rows.items():
            c = cur.get((city, d))
            if not c:
                continue
            for i in range(6):
                if t[i] is not None and c[i] is not None and t[i] != c[i]:
                    changed += 1
                    per_city_max[city] = max(per_city_max.get(city, 0), abs(t[i] - c[i]))
    print(f'\n  cells changed vs current CSV: {changed}  (calc->official, large is OK)')
    print('  per-city max delta (min):')
    for city in sorted(per_city_max, key=lambda c: -per_city_max[c]):
        print(f'    {city:22} {per_city_max[city]}')

    # seam diagnostic at both gap edges for the most extreme cities
    print('\n  gap seams (06/05->06/06 and 06/15->06/16):')
    for city in ['Adrar', 'Tindouf', 'Illizi', 'In Guezzam', 'Reggane']:
        rows = adrar_rows if city == 'Adrar' else city_rows[city]
        d5, d6 = date(2026, 6, 5), date(2026, 6, 6)
        d15, d16 = date(2026, 6, 15), date(2026, 6, 16)
        j1 = max(abs(rows[d6][i] - rows[d5][i]) for i in range(6))
        j2 = max(abs(rows[d16][i] - rows[d15][i]) for i in range(6))
        print(f'    {city:20} 05->06 maxjump={j1:2d}   15->16 maxjump={j2:2d}')

    if issues:
        print('\n!! structural issues present — NOT writing CSV. Inspect first.')
        return
    if not write:
        print('\n[dry-run] add --write to overwrite the CSV.')
        return

    # ── WRITE CSV (UPD cities fully replaced with 365 days each) ───────────────
    keep = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        rd = csv.reader(f)
        header = next(rd)
        for r in rd:
            if r[0] not in UPD:
                keep.append(r)

    new = []

    def emit(city, rows):
        for d, t in sorted(rows.items()):
            new.append([city, f'{d.day:02d}/{d.month:02d}/{d.year}']
                       + [pstr(x) for x in t])

    emit('Adrar', adrar_rows)
    for city, rows in city_rows.items():
        emit(city, rows)

    from collections import Counter
    cnt = Counter(r[0] for r in new)
    bad = {c: n for c, n in cnt.items() if n != 365}
    if bad:
        print('\n!! row-count mismatch (expected 365):', bad, '— NOT writing.')
        return

    out = keep + new

    def sk(r):
        dd, mm, yy = r[1].split('/')
        return (r[0], date(int(yy), int(mm), int(dd)))

    out.sort(key=sk)
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        wr = csv.writer(f)
        wr.writerow(header)
        wr.writerows(out)
    print(f'\nWrote CSV: kept={len(keep)} new={len(new)} total={len(out)}')


if __name__ == '__main__':
    main()
