"""
Rebuild Algiers (base) + its 25 surrounding wilayas from the official 1448 booklet:
  - Algiers_PrayerTimes_1448.xlsx       (official Algiers base, continuous Hijri
                                          year 1448: 355 days 2026/06/16..2027/06/05,
                                          cols: date|day|hijri_day|fajr|sunrise|
                                          dhuhr|asr|maghrib|isha, 12 month blocks)
  - Algiers_TimeDifferences_1448.xlsx   (official offsets, WIDE: 12 months x 6
                                          prayers x 25 cities, constant column order)

city[date] = Algiers_base[date] + offset[hijri_month_of_date][city][prayer]

Date -> calendar-2026 mapping (the CSV is a single reusable 2026 calendar):
  2026/01/01..06/05 -> xlsx 2027MMDD rows (same MM/DD, year-agnostic <1 min drift)
  2026/06/06..06/15 -> GAP (file covers neither year): keep current CSV values
  2026/06/16..12/31 -> xlsx 2026MMDD rows
Maghrib base already carries the official +3; offset applies on top (no extra +3).

Overwrites only Algiers + the 25 offset cities in assets/csv/algeria_prayer_times_2026.csv.
Dry-run by default (prints validation + diff). Pass --write to persist.
Run:  python tool/rebuild_algiers_from_1448_xlsx.py [--write]
"""
import sys, io, os, re, csv, unicodedata
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH  = os.path.join(BASE, 'assets', 'csv', 'algeria_prayer_times_2026.csv')
BASE_XLSX = os.path.join(BASE, 'Algiers_PrayerTimes_1448.xlsx')
OFF_XLSX  = os.path.join(BASE, 'Algiers_TimeDifferences_1448.xlsx')


def clean(s):
    # strip Arabic diacritics (Unicode nonspacing marks) — letters preserved.
    # NEVER use a regex char-class for this (the ؚ-ٰ range swallows letters).
    s = ''.join(c for c in (s or '') if unicodedata.category(c) != 'Mn').strip()
    return s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')


PRAYERS = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
PRAYER_AR = {clean(k): v for k, v in {
    'الفجر': 'fajr', 'الشروق': 'sunrise', 'الظهر': 'dhuhr',
    'العصر': 'asr', 'المغرب': 'maghrib', 'العشاء': 'isha',
}.items()}

MONTHS = [clean(m) for m in [
    'محرم', 'صفر', 'ربيع الاول', 'ربيع الثاني', 'جمادى الاولى', 'جمادى الثانية',
    'رجب', 'شعبان', 'رمضان', 'شوال', 'ذو القعدة', 'ذو الحجة',
]]


def month_key(s):
    c = re.sub(r'[0-9٠-٩]+', '', clean(s)).strip()
    return c if c in MONTHS else None


# 25 Algiers-booklet cities: Arabic offset header -> CSV English name (all exist).
ALGIERS_AR_TO_EN = {
    'الطارف': 'El Tarf', 'سوق أهراس': 'Souk Ahras', 'عنابة': 'Annaba',
    'قالمة': 'Qalma', 'أم البواقي': 'Um al-Bouaki', 'سكيكدة': 'Skikda',
    'قسنطينة': 'Constantine', 'ميلة': 'Milah', 'جيجل': 'Jijel', 'سطيف': 'Setif',
    'بجاية': 'Bjaiya', 'برج بوعريريج': 'Bordj Bou Arreridj', 'المسيلة': 'Msila',
    'تيزي وزو': 'Tizi Ouzou', 'دلس': 'Dellys', 'البويرة': 'Bouira',
    'بومرداس': 'Boumerdes', 'البليدة': 'Blida', 'المدية': 'Medea',
    'تيبازة': 'Tipaza', 'عين الدفلى': 'Ain Defla', 'الشلف': 'Chlef',
    'غليزان': 'Relizane', 'مستغانم': 'Mostaganem', 'وهران': 'Oran',
}
AR_CLEAN = {clean(k): v for k, v in ALGIERS_AR_TO_EN.items()}


def ptime(s):
    if s is None:
        return None
    m = re.match(r'(\d{1,2}):(\d{2})', str(s).strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def pstr(m):
    return '' if m is None else f'{m // 60:02d}:{m % 60:02d}'


# ── read Algiers base (continuous 1448 Hijri year) ────────────────────────────
def read_base():
    ws = openpyxl.load_workbook(BASE_XLSX, data_only=True).active
    times, hijri = {}, {}
    cur_month = None
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        s = str(a).strip()
        m = re.match(r'(\d{4})/(\d{2})/(\d{2})', s)
        if m:
            key = m.group(1) + m.group(2) + m.group(3)
            vals = [ptime(row[3 + i]) for i in range(6)]
            if vals[0] is None:
                continue
            times[key] = dict(zip(PRAYERS, vals))
            hijri[key] = cur_month
        elif s.startswith('التاريخ'):
            continue  # repeated column header
        else:
            mk = month_key(s)
            if mk:
                cur_month = mk
    return times, hijri


# ── read offsets (wide: one block per Hijri month) ────────────────────────────
def read_offsets():
    ws = openpyxl.load_workbook(OFF_XLSX, data_only=True).active
    off = {}            # city_en -> month -> prayer -> int
    cur_month = None
    col_city = None     # {col_index: city_en}
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        s = clean(str(a))
        if s.startswith('الصلاة'):  # header row -> capture the 25 city columns
            col_city = {ci: AR_CLEAN.get(clean(str(row[ci]))) if row[ci] else None
                        for ci in range(1, 26)}
            continue
        pk = PRAYER_AR.get(s)
        if pk and cur_month and col_city:
            for ci, city in col_city.items():
                if city is None or row[ci] is None:
                    continue
                try:
                    off.setdefault(city, {}).setdefault(cur_month, {})[pk] = \
                        int(round(float(row[ci])))
                except (ValueError, TypeError):
                    continue
            continue
        mk = month_key(s)
        if mk:
            cur_month = mk
    return off


# ── build calendar-2026 for the Algiers base ──────────────────────────────────
def build_cal(times, hijri):
    cal = {}
    d = date(2026, 1, 1)
    while d <= date(2026, 12, 31):
        if d >= date(2026, 6, 16):
            k = f'2026{d.month:02d}{d.day:02d}'
        elif d <= date(2026, 6, 5):
            k = f'2027{d.month:02d}{d.day:02d}'
        else:                       # 06/06..06/15 — uncovered by the booklet
            cal[d] = None
            d += timedelta(days=1)
            continue
        row = times.get(k)
        cal[d] = ({**row, 'hijri': hijri.get(k, '')} if row else None)
        d += timedelta(days=1)
    return cal


GAP = {date(2026, 6, d) for d in range(6, 16)}          # 06/06..06/15 keep-current
UPD = set(ALGIERS_AR_TO_EN.values()) | {'Algiers'}
LABELS = ['Fajr', 'Sunrise', 'Dhuhr', 'Asr', 'Maghrib', 'Isha']


def main():
    write = '--write' in sys.argv
    print('Reading base + offsets...')
    times, hijri = read_base()
    off = read_offsets()
    print(f'  base days={len(times)} (first={min(times)} last={max(times)})  '
          f'offset cities={len(off)}/25')

    missing = [c for c in ALGIERS_AR_TO_EN.values() if c not in off]
    if missing:
        print('  !! MISSING offset cities:', missing)
    for c, mo in off.items():
        bad = [m for m, pp in mo.items() if len(pp) != 6]
        if len(mo) != 12 or bad:
            print(f'  !! {c}: months={len(mo)} incomplete={bad}')

    cal = build_cal(times, hijri)
    filled = sum(1 for v in cal.values() if v)
    print(f'  calendar 2026: {filled}/365 filled, gap={365 - filled} '
          f'({sorted(d.strftime("%d/%m") for d, v in cal.items() if v is None)})')

    # build city rows: city_en -> {date: [6 mins]}
    city_rows = {}
    for city in ALGIERS_AR_TO_EN.values():
        rows = {}
        for d, e in cal.items():
            if e is None:
                continue
            mo = off.get(city, {}).get(e['hijri'], {})
            rows[d] = [e[p] + mo.get(p, 0) for p in PRAYERS]
        city_rows[city] = rows
    alg_rows = {d: [e[p] for p in PRAYERS] for d, e in cal.items() if e}

    # ── VALIDATION ────────────────────────────────────────────────────────────
    print('\n── validation ──')
    issues = 0
    all_rows_for_check = list(city_rows.items()) + [('Algiers', alg_rows)]
    # 1) strict monotonic ordering each day
    for city, rows in all_rows_for_check:
        for d, t in rows.items():
            if any(x is None for x in t):
                continue
            if not (t[0] < t[1] < t[2] < t[3] < t[4] < t[5]):
                print(f'  ORDER  {city} {d}: {[pstr(x) for x in t]}')
                issues += 1
    # 2) day-to-day smoothness per prayer (catches offset typos at month seams)
    TH = [12, 12, 8, 12, 10, 14]
    for city, rows in all_rows_for_check:
        ds = sorted(rows.keys())
        for a, b in zip(ds, ds[1:]):
            if (b - a).days != 1:
                continue              # skip across the kept-current gap
            for i in range(6):
                j = rows[b][i] - rows[a][i]
                if abs(j) > TH[i]:
                    print(f'  JUMP   {city} {LABELS[i]} {a}->{b}: '
                          f'{pstr(rows[a][i])}->{pstr(rows[b][i])} ({j:+d})')
                    issues += 1
    print(f'  structural issues: {issues}')

    # 3) diff vs current CSV + seam check at the kept-current gap edges
    cur = {}
    with open(CSV_PATH, encoding='utf-8') as f:
        for r in csv.reader(f):
            if r[0] in UPD and '/' in r[1]:
                dd, mm, yy = r[1].split('/')
                cur[(r[0], date(int(yy), int(mm), int(dd)))] = [ptime(x) for x in r[2:8]]
    changed = maxch = 0
    samples = []
    for city, rows in all_rows_for_check:
        for d, t in rows.items():
            c = cur.get((city, d))
            if not c:
                continue
            for i in range(6):
                if t[i] is not None and c[i] is not None and t[i] != c[i]:
                    changed += 1
                    maxch = max(maxch, abs(t[i] - c[i]))
                    if len(samples) < 14:
                        samples.append(f'  {city} {d} {LABELS[i]}: '
                                       f'{pstr(c[i])}->{pstr(t[i])} ({t[i] - c[i]:+d})')
    print(f'\n  cells changed vs current CSV: {changed}  (max delta {maxch} min)')
    for sline in samples:
        print(sline)

    # seam: new 06/05 -> kept 06/06, and kept 06/15 -> new 06/16 (Algiers)
    print('\n  gap seam (Algiers, new vs kept-current):')
    for d in (date(2026, 6, 5), date(2026, 6, 6), date(2026, 6, 15), date(2026, 6, 16)):
        src = 'kept' if d in GAP else 'new '
        t = alg_rows.get(d) or cur.get(('Algiers', d))
        print(f'    {d} [{src}] {[pstr(x) for x in t]}')

    if issues:
        print('\n!! structural issues present — NOT writing CSV. Inspect first.')
        return
    if not write:
        print('\n[dry-run] add --write to overwrite the CSV.')
        return

    # ── WRITE CSV ─────────────────────────────────────────────────────────────
    keep = []
    with open(CSV_PATH, encoding='utf-8', newline='') as f:
        rd = csv.reader(f)
        header = next(rd)
        for r in rd:
            if r[0] not in UPD:
                keep.append(r)
                continue
            dd, mm, yy = r[1].split('/')
            if date(int(yy), int(mm), int(dd)) in GAP:
                keep.append(r)        # keep-current for the uncovered gap days

    new = []

    def emit(city, rows):
        for d, t in sorted(rows.items()):
            new.append([city, f'{d.day:02d}/{d.month:02d}/{d.year}']
                       + [pstr(x) for x in t])

    emit('Algiers', alg_rows)
    for city, rows in city_rows.items():
        emit(city, rows)

    # assert every updated city is back to exactly 365 rows
    from collections import Counter
    cnt = Counter(r[0] for r in keep + new if r[0] in UPD)
    bad = {c: n for c, n in cnt.items() if n != 365}
    if bad:
        print('\n!! row-count mismatch (expected 365):', bad, '— NOT writing.')
        return

    all_rows = keep + new

    def sk(r):
        dd, mm, yy = r[1].split('/')
        return (r[0], date(int(yy), int(mm), int(dd)))

    all_rows.sort(key=sk)
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        wr = csv.writer(f)
        wr.writerow(header)
        wr.writerows(all_rows)
    print(f'\nWrote CSV: kept={len(keep)} new={len(new)} total={len(all_rows)}')


if __name__ == '__main__':
    main()
